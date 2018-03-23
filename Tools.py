# coding=utf-8
from openpyxl.styles import Border
import math


def restructure_border(border, top=None, left=None, right=None, bottom=None):
    top = border.top if top is None else top.top
    left = border.left if left is None else left.left
    right = border.right if right is None else right.right
    bottom = border.bottom if bottom is None else bottom.bottom
    return Border(top=top, right=right, bottom=bottom, left=left)


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = restructure_border(cell.border, top=top)
    for cell in rows[-1]:
        cell.border = restructure_border(cell.border, bottom=bottom)

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = restructure_border(l.border, left=left)
        r.border = restructure_border(r.border, right=right)
        if fill:
            for c in row:
                c.fill = fill


def decimal2letter(x):
    res = ['@']
    for i in range(x):
        count = 0
        ex = 1
        while ex:
            tmp = ord(res[count])
            res[count] = chr(((tmp - 65 + 1) % 26) + 65)
            ex = (tmp - 65 + 1) / 26
            count += 1
            if count > len(res) - 1 and ex:
                res.append('@')
    return ''.join(res)[::-1]


def coordinate_transfer(x, y):
    return decimal2letter(y) + str(x)


def as_text(value):
    if value is None:
        return ""
    elif type(value) is int:
        return str(value)
    return str(value.encode('utf-8'))
