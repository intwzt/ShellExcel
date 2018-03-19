# coding=utf-8
from openpyxl import Workbook

from Common import bg_color, target_mapper, ref_mapper, formula_type, fill_pattern, side_pattern, \
    double_thin_top_border, thick_border
from Common import border_pattern, alignment_pattern, font_pattern, side_style, font_style, alignment
from Cube import Cube
from Tools import coordinate_transfer, style_range
from Style import Style


class Table:
    def __init__(self, ox, oy):
        # add title row
        ox += 1
        self.owner = None
        self.follower = []
        self.header = None
        self.origin = [ox, oy]
        self.table_body = [ox, oy]
        self.title = None

        self.wb = Workbook()
        self.ws = self.wb.active
        self.end_point = [ox, oy]

        # remove GridLines
        self.ws.sheet_view.showGridLines = False

    def _cal_end_point(self):
        body_len_x = 0
        body_len_y = 0
        # cal body_len_y based on table body
        if len(self.follower) > 0:
            # get follower column parameter
            follower_target_column_gap = len(self.follower[0].target)
            follower_ref_column_gap = len(self.follower[0].ref)
            follower_column_gap = follower_target_column_gap + follower_ref_column_gap
            body_len_y += follower_column_gap * len(self.follower)
        # get owner column parameter
        owner_column_gap = len(self.owner.target)
        body_len_y += owner_column_gap * 1

        # cal body_len_x based on table header
        if self.header is not None:
            body_len_x += self.header.x
        self.end_point = [body_len_x + self.origin[0] - 1, body_len_y + self.origin[1] + 1]

    def _add_table_border(self):
        coordinate = coordinate_transfer(self.origin[0], self.origin[1]) + ':' + \
                     coordinate_transfer(self.end_point[0], self.end_point[1])
        style_range(self.ws, coordinate, thick_border)

    def assign_header(self, header):
        self.header = header
        self.table_body = [self.table_body[0], self.table_body[1] + header.y]

    def assign_title(self, title):
        self.title = title

    def _render_table_title(self):
        style = Style(bg_color[4], border=None, font=font_style[3], al=alignment[4])
        self._write_cube_to_book(self.origin[0] - 1, self.origin[1], Cube(bg_color[4], value=self.title, style=style))

    def _check_cube_style(self, x, y):
        for item in self.header.merge:
            if item['coordinate'][0] == x and item['coordinate'][1] == y:
                return item['style']
        return None

    def _render_table_header(self):
        # render cell with style
        for i in range(self.header.x):
            for j in range(self.header.y):
                x = i + self.origin[0] - 1
                # y should add title
                y = j + self.origin[1] - 1
                current_value = self.header.matrix[i][j]
                checked_style = self._check_cube_style(i, j)
                current_cube = Cube(bg_color[4], value=current_value) if checked_style is None else Cube(
                    checked_style.fill, value=current_value)
                self._write_cube_to_book(x + 1, y + 1, current_cube)

        # merge cell
        if self.header.merge is not None:
            for m in self.header.merge:
                start_row = m['coordinate'][0] + self.origin[0]
                start_column = m['coordinate'][1] + self.origin[1]
                end_row = m['coordinate'][2] + self.origin[0]
                end_column = m['coordinate'][3] + self.origin[1]
                self.ws.merge_cells(start_row=start_row,
                                    start_column=start_column,
                                    end_row=end_row,
                                    end_column=end_column)
                # get style of cell need merged
                current_style = m['style']

                self._merge_body_cell(end_column, end_row, start_column, start_row,
                                      border_pattern[current_style.border], fill_pattern[current_style.fill],
                                      font_pattern[current_style.font], alignment_pattern[current_style.al])
                # set border
                coordinate = coordinate_transfer(start_row, start_column) + ':' + coordinate_transfer(end_row, end_column)
                style_range(self.ws, coordinate,
                            border=border_pattern[current_style.border],
                            fill=fill_pattern[current_style.fill],
                            font=font_pattern[current_style.font],
                            alignment=alignment_pattern[current_style.al])

    def assign_person(self, owner, follower=None):
        self.owner = owner
        if follower:
            for f in follower:
                self.follower.append(f)

    def _set_formula(self, x, y, formula):
        self.ws[coordinate_transfer(x, y)] = formula


    def _style_factory(self, c, cube):
        border = None if cube.style.border is None else border_pattern[cube.style.border]
        fill = None if cube.style.fill is None else fill_pattern[cube.style.fill]
        font = None if cube.style.font is None else font_pattern[cube.style.font]
        al = None if cube.style.al is None else alignment_pattern[cube.style.al]
        c.border = border
        c.fill = fill
        c.font = font
        c.alignment = al

    def _write_cube_to_book(self, x, y, cube):
        c = self.ws.cell(x, y, cube.value)
        if cube.formula is not None:
            self._set_formula(x, y, cube.formula)
        # write style
        # get cell range
        self._style_factory(c, cube)

    def print_info(self):
        self.owner.print_info()

    def save_workbook(self):
        self.wb.save('result.xlsx')

    def render(self):
        self._render_table_title()
        # render header
        self._render_table_header()
        # render body
        index_x = self.table_body[0]
        index_y = self.table_body[1]
        counter = 0
        num_of_column = 0 if self.owner is None else len(self.owner.target[target_mapper[0]].container)

        self._render_owner(counter, index_x, index_y, num_of_column)
        # modify index_x and reset counter
        index_y += 3
        counter = 0

        num_of_column = 0 if len(self.follower) is None else len(self.follower[0].target[target_mapper[0]].container)
        self._render_follower(counter, index_x, index_y, num_of_column)

        # add table border
        self._cal_end_point()
        self._add_table_border()

    def _render_follower(self, counter, index_x, index_y, num_of_column):
        # write follower if exist
        if len(self.follower) > 0:
            # for all follower
            for f in range(len(self.follower)):

                # get some column parameter
                person_target_column_number = len(self.follower[f].target)
                person_ref_column_number = len(self.follower[f].ref)
                person_column_number = person_target_column_number + person_ref_column_number

                inner_counter = 0
                # write person name
                for i in range(person_column_number):
                    self._write_cube_to_book(index_x + counter + inner_counter, index_y + i + f * person_column_number,
                                             Cube(bg_color[4], self.follower[f].name))

                # merge name cells
                start_row = index_x + counter + inner_counter
                start_column = index_y + f * person_column_number
                end_row = index_x + counter + inner_counter
                end_column = index_y + f * person_column_number + person_column_number - 1

                self._merge_body_cell(end_column, end_row, start_column, start_row,
                                      border_pattern[side_style[1]], fill_pattern[bg_color[4]],
                                      font_pattern[font_style[2]], alignment_pattern[alignment[2]])

                inner_counter += 1
                # write column name
                col_name_style = Style(bg_color[4], font=font_style[2], al=alignment[2])
                for i in range(person_target_column_number):
                    self._write_cube_to_book(index_x + counter + inner_counter, index_y + i + f * person_column_number,
                                             Cube(bg_color[4], ref_mapper[i], style=col_name_style))
                for i in range(person_target_column_number, person_column_number):
                    self._write_cube_to_book(index_x + counter + inner_counter, index_y + i + f * person_column_number,
                                             Cube(bg_color[4], target_mapper[i - person_target_column_number],
                                                  style=col_name_style))

                inner_counter += 1
                # write column data
                for i in range(num_of_column):
                    for j in range(person_ref_column_number):
                        # cast formula type to real formula if the cube has formula
                        if formula_type[1] == self.follower[f].ref[ref_mapper[j]].container[i].formula:
                            start_cube = coordinate_transfer(index_x + counter + inner_counter,
                                                             index_y + j + f * person_column_number)
                            end_cube = coordinate_transfer(index_x + counter + i + inner_counter - 1,
                                                           index_y + j + f * person_column_number)
                            formula = "=SUM(" + start_cube + ":" + end_cube + ")"
                            self.follower[f].ref[ref_mapper[j]].container[i].set_formula(formula)
                        self._write_cube_to_book(index_x + counter + i + inner_counter,
                                                 index_y + j + f * person_column_number,
                                                 self.follower[f].ref[ref_mapper[j]].container[i])
                for i in range(num_of_column):
                    for j in range(person_target_column_number):
                        # cast formula type to real formula if the cube has formula
                        if formula_type[1] == self.follower[f].target[target_mapper[j]].container[i].formula:
                            start_cube = coordinate_transfer(index_x + counter + inner_counter,
                                                             index_y + j + person_ref_column_number + f * person_column_number)
                            end_cube = coordinate_transfer(index_x + counter + i + inner_counter - 1,
                                                           index_y + j + person_ref_column_number + f * person_column_number)
                            formula = "=SUM(" + start_cube + ":" + end_cube + ")"
                            self.follower[f].target[target_mapper[j]].container[i].set_formula(formula)
                        self._write_cube_to_book(index_x + counter + i + inner_counter, index_y + j +
                                                 person_ref_column_number + f * person_column_number,
                                                 self.follower[f].target[target_mapper[j]].container[i])
            counter += 1

    def _render_owner(self, counter, index_x, index_y, num_of_column):
        # write owner if exist
        if self.owner:
            # get some column parameter
            person_target_column_number = len(self.owner.target)
            person_column_number = person_target_column_number
            # write person name
            for i in range(person_column_number):
                self._write_cube_to_book(index_x + counter, index_y + i,
                                         Cube(bg_color[4], self.owner.name))

            # merge name cells
            start_row = index_x + counter
            start_column = index_y
            end_row = index_x + counter
            end_column = index_y + person_column_number - 1
            self._merge_body_cell(end_column, end_row, start_column, start_row,
                                  border_pattern[side_style[1]], fill_pattern[bg_color[4]],
                                  font_pattern[font_style[2]], alignment_pattern[alignment[2]])

            counter += 1
            # write column name
            col_name_style = Style(bg_color[4], font=font_style[2], al=alignment[2])
            for i in range(person_column_number):
                self._write_cube_to_book(index_x + counter, index_y + i,
                                         Cube(bg_color[4], target_mapper[i], style=col_name_style))
            counter += 1
            # write column data
            for i in range(num_of_column):
                for j in range(person_target_column_number):
                    # cast formula type to real formula if the cube has formula
                    if formula_type[1] == self.owner.target[target_mapper[j]].container[i].formula:
                        start_cube = coordinate_transfer(index_x + counter, index_y + j)
                        end_cube = coordinate_transfer(index_x + counter + i - 1, index_y + j)
                        formula = "=SUM(" + start_cube + ":" + end_cube + ")"
                        self.owner.target[target_mapper[j]].container[i].set_formula(formula)
                    self._write_cube_to_book(index_x + counter + i, index_y + j,
                                             self.owner.target[target_mapper[j]].container[i])

    def _merge_body_cell(self, end_column, end_row, start_column, start_row,
                         border, fill, font, al):
        self.ws.merge_cells(start_row=start_row,
                            start_column=start_column,
                            end_row=end_row,
                            end_column=end_column)
        # set merge cells style
        coordinate = coordinate_transfer(start_row, start_column) + ':' + coordinate_transfer(end_row, end_column)
        style_range(self.ws, coordinate, border, fill, font, al)
