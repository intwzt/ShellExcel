# coding=utf-8
from PageParser import PageParser
from openpyxl import load_workbook

from template.SheetType import sheet_type


class Parser:
    def __init__(self, filename, table_type):
        self.filename = filename
        self.table_type = table_type
        self.wb = None
        self.load_workbook()

    def load_page(self, page):
        return self.wb.worksheets[page]

    def load_workbook(self):
        self.wb = load_workbook(self.filename, data_only=True, read_only=True)

    def parse(self, page=-1):
        if page == 1:
            return 'User Instruction Page'
        elif page == 0:
            return 'Page Number Error'
        elif page > 1:
            current_ws = self.load_page(page-1)
            page_parser = PageParser(current_ws, self.table_type)
            return {1: page_parser.parse_attached_table(), 2: page_parser.parse_main_table()}
        else:
            num_of_page = len(sheet_type[self.table_type]['page'])
            data = {}
            for i in range(1, num_of_page):
                current_ws = self.load_page(i-1)
                page_parser = PageParser(current_ws, self.table_type)
                page_name = sheet_type[self.table_type]['page'][i]
                data[page_name] = {1: page_parser.parse_attached_table(), 2: page_parser.parse_main_table()}
            return data
