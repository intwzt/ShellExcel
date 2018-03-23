# coding=utf-8
import os

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

from excel_generator.Sheet import Sheet
from template.SheetType import sheet_type, RSM


class SheetBook:
    def __init__(self, table_type, filename, data):
        self.wb = Workbook()
        self.table_type = table_type
        self.filename = filename
        self.data = data

        self._copy_from_template()

        self.wb = load_workbook('../excel_files/' + self.filename)

        # add image for user instruction
        self._insert_user_instruction_image()

        # get page info from SheetType
        self.num_of_page = len(sheet_type[self.table_type]['page'])
        self.pages = []

        # create pages
        self._create_pages()

        # remove GridLines
        for page in self.pages:
            page.sheet_view.showGridLines = False

        # render sheet
        self._render_sheet()

    def _resize_image(self, img, height, width):
        img.height = height
        img.width = width
        return img

    def _insert_user_instruction_image(self):
        image1_coordinate = 'E3'
        image2_coordinate = 'E35'
        page_name = self.wb.sheetnames[0]
        img1 = Image('../excel_files/image/1.png')
        img2 = Image('../excel_files/image/2.png')
        self._resize_image(img1, 600, 1400)
        self._resize_image(img2, 800, 1400)
        self.wb[page_name].add_image(img1, image1_coordinate)
        self.wb[page_name].add_image(img2, image2_coordinate)

    def _render_sheet(self):
        for i in range(0, self.num_of_page - 1):
            page_name = sheet_type[self.table_type]['page'][i+1]
            sheet = Sheet(self.pages[i], i+1, self.data[page_name], RSM)
            sheet.render()

    def _create_pages(self):
        for i in range(1, self.num_of_page):
            self.pages.append(self.wb.create_sheet(sheet_type[self.table_type]['page'][i], i))

    def _copy_from_template(self):
        command = 'cp ../excel_files/template.xlsx ../excel_files/' + self.filename
        os.system(command)

    def save_book(self):
        self.wb.save('../excel_files/' + self.filename)
