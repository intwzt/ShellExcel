# coding=utf-8
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font

role = {1: 'owner', 2: 'follower'}

bg_color = {1: 'gray', 2: 'yellow', 3: 'blue', 4: 'white'}

fill_pattern = {bg_color[1]: PatternFill('solid', fgColor="e6e6e6"),
                bg_color[2]: PatternFill('solid', fgColor='fff1ce'),
                bg_color[3]: PatternFill('solid', fgColor='4674c1'),
                bg_color[4]: PatternFill('solid', fgColor='ffffff')}

font_color = {1: 'red'}

formula = {1: 'None', 2: 'have'}

target_mapper = ['Target UC3 $', 'Target UNP $', 'Target Portfolio %']

ref_mapper = ['Ref UC3 $', 'Ref UNP $', 'Ref Portfolio %']

attach_column_type = ['LE KL', 'Market size KL',
                      'Market Share %', 'Target Volume KL',
                      'Target C3 $', 'Target Proceed $']

column_type = {1: 'target', 2: 'ref'}

side_style = {1: 'thin', 2: 'thick', 3: 'double_thin_top'}
side_pattern = {side_style[1]: Side(border_style="thin", color="8e8d8d"),
                side_style[2]: Side(border_style="thick", color="000000"),
                side_style[3]: Side(border_style="double", color="000000"),
                }
thin_border = Border(top=side_pattern[side_style[1]], left=side_pattern[side_style[1]],
                     right=side_pattern[side_style[1]], bottom=side_pattern[side_style[1]])
thick_border = Border(top=side_pattern[side_style[2]], left=side_pattern[side_style[2]],
                      right=side_pattern[side_style[2]], bottom=side_pattern[side_style[2]])
double_thin_top_border = Border(top=side_pattern[side_style[3]], left=side_pattern[side_style[1]],
                                right=side_pattern[side_style[1]], bottom=side_pattern[side_style[1]])
border_pattern = {side_style[1]: thin_border, side_style[2]: thick_border, side_style[3]: double_thin_top_border}

alignment = {1: 'left', 2: 'center', 3: 'right', 4: 'title', 5: 'header'}
alignment_pattern = {alignment[1]: Alignment(horizontal="left", vertical="center", wrap_text=True),
                     alignment[2]: Alignment(horizontal="center", vertical="center", wrap_text=True),
                     alignment[3]: Alignment(horizontal="right", vertical="center", wrap_text=True),
                     alignment[4]: Alignment(horizontal="left"),
                     alignment[5]: Alignment(horizontal="left", vertical="bottom")}

column_last_row = {1: 'null', 2: 'hold', 3: 'cal', 4: 'negative'}

formula_type = {1: 'column_total', 2: 'inner_product', 3: 'null'}

font_style = {1: 'normal', 2: 'bold', 3: 'table_title'}
font_pattern = {font_style[1]: Font(b=False, name="Microsoft YaHei", color="000000", size=10),
                font_style[2]: Font(b=True, name="Microsoft YaHei", color="000000", size=10),
                font_style[3]: Font(b=True, name="Microsoft YaHei", color="000000", size=14)}
