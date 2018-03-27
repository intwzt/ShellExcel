# coding=utf-8
from Tools import coordinate_transfer, style_range, divide_column
from excel_generator.Common import bg_color, font_style, alignment, border_pattern, fill_pattern, alignment_pattern, \
    font_pattern, side_style, thick_border
from excel_generator.Cube import Cube
from excel_generator.Style import Style
from excel_generator.mock.SGM import page2
from excel_generator.tool.RenderCore import RenderCore
from template.HeaderTemplate import common_header_sgm_1


class MainTable:
    def __init__(self, core, title, header, data, x, y):
        self.title = title
        self.header = header
        self.data = data
        self.origin = [x, y]
        self.end_point = [x, y]
        self.core = core

    def _adjust_column_width(self, width):
        for column_cells in self.core.ws.columns:
            # length = max(len(as_text(cell.value)) for cell in column_cells)
            self.core.ws.column_dimensions[column_cells[0].column].width = width

    def _render_title(self):
        style = Style(bg_color[4], border=None, font=font_style[3], al=alignment[4])

        self.core.write_cube_to_book(self.origin[0], self.origin[1],
                                     Cube(bg_color[4], value=self.title, style=style))

    def _check_cube_style(self, x, y):
        for item in self.header['merge']:
            if item['coordinate'][0] == x and item['coordinate'][1] == y:
                return item['style']
        return Style(bg_color[4], border=side_style[1], font=font_style[2], al=alignment[2])

    def _merge_body_cell(self, end_column, end_row, start_column, start_row,
                         border, fill, font, al):
        self.core.ws.merge_cells(start_row=start_row,
                                 start_column=start_column,
                                 end_row=end_row,
                                 end_column=end_column)
        # set merge cells style
        coordinate = coordinate_transfer(start_row, start_column) + ':' + coordinate_transfer(end_row, end_column)
        style_range(self.core.ws, coordinate, border, fill, font, al)

    def _render_header(self):
        # render cell with style
        scale_x = self.header['scale'][0]
        scale_y = self.header['scale'][1]
        for i in range(scale_x):
            for j in range(scale_y):
                x = i + self.origin[0] + 1
                # y should add title
                y = j + self.origin[1]
                current_value = self.header['data'][i][j]
                checked_style = self._check_cube_style(i, j)
                current_cube = Cube(style=checked_style, value=current_value)
                self.core.write_cube_to_book(x, y, current_cube)

        # merge cell
        if self.header['merge'] is not None:
            bx = self.origin[0] + 1
            by = self.origin[1]
            for m in self.header['merge']:
                start_row = m['coordinate'][0] + bx
                start_column = m['coordinate'][1] + by
                end_row = m['coordinate'][2] + bx
                end_column = m['coordinate'][3] + by
                self.core.ws.merge_cells(start_row=start_row,
                                         start_column=start_column,
                                         end_row=end_row,
                                         end_column=end_column)
                # get style of cell need merged
                current_style = m['style']

                self._merge_body_cell(end_column, end_row, start_column, start_row,
                                      border_pattern[current_style.border], fill_pattern[current_style.fill],
                                      font_pattern[current_style.font], alignment_pattern[current_style.al])
                # set border
                coordinate = coordinate_transfer(start_row, start_column) + ':' + coordinate_transfer(end_row,
                                                                                                      end_column)
                style_range(self.core.ws, coordinate,
                            border=border_pattern[current_style.border],
                            fill=fill_pattern[current_style.fill],
                            font=font_pattern[current_style.font],
                            alignment=alignment_pattern[current_style.al])

    def _list_divide(self, l1, l2):
        res = []
        for i in range(len(l1)):
            res.append(float(l1[i]) / float(l2[i]))
        return res

    def _list_divide_subtract_1(self, l1, l2):
        res = []
        for i in range(len(l1)):
            res.append(float(l1[i]) / float(l2[i]) - 1.0)
        return res

    def _cal_end_point(self):
        hx = self.header['scale'][0]
        hy = self.header['scale'][1]
        scale_x = len(self.data)
        self.end_point = [self.origin[0] + hx + scale_x + 1, self.origin[1] + hy - 1]

    def _add_table_border(self):
        coordinate = coordinate_transfer(self.origin[0] + 1, self.origin[1]) + ':' + \
                     coordinate_transfer(self.end_point[0], self.end_point[1])
        style_range(self.core.ws, coordinate, thick_border)

    def _get_all_data_by_column(self, name):
        res = []
        if name == 'LE KL':
            for p in self.data:
                res.append(p['LE KL'])
            return res
        elif name == 'Market size KL(this year)':
            for p in self.data:
                res.append(p['Market size KL(this year)'])
            return res
        elif name == 'Market size KL(last year)':
            for p in self.data:
                res.append(p['Market size KL(last year)'])
            return res
        elif name == 'Market Share %':
            return self._list_divide(self._get_all_data_by_column('LE KL'),
                                     self._get_all_data_by_column('Market size KL(this year)'))
        elif name == 'Market Growth %':
            return self._list_divide_subtract_1(self._get_all_data_by_column('Market size KL(this year)'),
                                                self._get_all_data_by_column('Market size KL(last year)'))
        else:
            pass

    def _formula_converter(self, name, row, flag, x, y, number_format):
        tmp_y = y - self.origin[1]
        if tmp_y < 4:
            style = Style(bg_color=bg_color[4], border=side_style[1], al=alignment[1])
        elif tmp_y >= 3 and flag == 0:
            style = Style(bg_color=bg_color[2], border=side_style[1], al=alignment[2])
        elif flag < 0:
            style = Style(bg_color=bg_color[4], border=side_style[1], al=alignment[2])
        else:
            style = Style(bg_color=bg_color[1], border=side_style[1])
        formula = ''
        if flag == 0:
            return Cube(style=style, number_format=number_format, value=row[name])
        elif name == 'Market Share %':
            formula = '={0}/{1}'.format(coordinate_transfer(x, y - 2), coordinate_transfer(x, y - 1))

        elif name == 'Market Growth %':
            formula = '={0}/{1}-1'.format(coordinate_transfer(x, y - 3), coordinate_transfer(x, y - 1))

        elif name == 'Market Share Score':
            a, b, c, d = divide_column(self._get_all_data_by_column('Market Share %'))
            formula = '=IF({0}<{1}, 5, IF({0}<{2}, 4, IF({0}<{3},3, IF({0}<{4}, 2, 1))))'.format(
                coordinate_transfer(x, y - 4), a, b, c, d)

        elif name == 'Market Growth Score':
            a, b, c, d = divide_column(self._get_all_data_by_column('Market Growth %'))
            formula = '=IF({0}<{1}, 5, IF({0}<{2}, 4, IF({0}<{3},3, IF({0}<{4}, 2, 1))))'.format(
                coordinate_transfer(x, y - 3), a, b, c, d)

        elif name == 'Market Share Score(0.75)':
            formula = '={0}*0.75'.format(coordinate_transfer(x, y - 3))

        elif name == 'Market Growth Score(0.15)':
            formula = '={0}*0.15'.format(coordinate_transfer(x, y - 3))

        elif name == 'Platform Score(0.1)':
            formula = '={0}*0.1'.format(coordinate_transfer(x, y - 3))

        elif name == 'Platform Score':
            formula = '={0}'.format(coordinate_transfer(x, y - 3))

        elif name == 'Total Score':
            formula = '=SUM({0}:{1})'.format(coordinate_transfer(x, y - 3), coordinate_transfer(x, y - 1))

        elif name == 'Increase %':
            row = len(self.data)
            formula = '={0}*({1}-{2})/SUMPRODUCT({3}:{4},{5}:{6})'.format(
                coordinate_transfer(x, y - 1),
                coordinate_transfer(self.origin[0] + 1 + row + 1, y + 1),
                coordinate_transfer(self.origin[0] + 1 + row, y - 13),
                coordinate_transfer(self.origin[0] + 1 + 1, y - 13),
                coordinate_transfer(self.origin[0] + row, y - 13),
                coordinate_transfer(self.origin[0] + 1 + 1, y - 1),
                coordinate_transfer(self.origin[0] + row, y - 1),
            )
        elif name == 'Ref Target KL':
            formula = '={0}*(1+{1})'.format(coordinate_transfer(x, y - 14), coordinate_transfer(x, y - 1))

        else:
            pass
        return Cube(style=style, formula=formula, number_format=number_format)

    def _render_body(self):
        bx = self.header['scale'][0] + self.origin[0] + 1
        by = self.origin[1]
        scale_x = len(self.data)
        scale_y = self.header['scale'][1]

        for i in range(scale_x):
            for j in range(scale_y):
                cx = bx + i
                cy = by + j
                col_name = self.header['data'][0][j]
                number_format = self.header['number_format'][j]
                formula_flag = self.header['formula'][j]
                current_cube = self._formula_converter(col_name, self.data[i], formula_flag, cx, cy, number_format)
                self.core.write_cube_to_book(cx, cy, current_cube)

    def _render_footer(self):
        # render total row
        sum_total = self.header['total']
        scale_x = len(self.data)
        foot_x = self.origin[0] + self.header['scale'][0] + scale_x
        foot_y = self.origin[1]
        for i in range(self.header['scale'][1]):
            if i < 3:
                current_cube = Cube(
                    style=Style(bg_color=bg_color[4], border=side_style[3], al=alignment[1], font=font_style[2]),
                    value='Sum Total')
                self.core.write_cube_to_book(foot_x, foot_y + i, current_cube)
            else:
                flag = sum_total[i]
                if flag == 1:
                    current_cube = Cube(
                        style=Style(bg_color=bg_color[1], border=side_style[3], al=alignment[3], font=font_style[2]))
                    start = coordinate_transfer(self.origin[0] + 1 + self.header['scale'][0], foot_y + i)
                    end = coordinate_transfer(foot_x - 1, foot_y + i)
                    current_cube.set_formula('=SUM({0}:{1})'.format(start, end))
                    self.core.write_cube_to_book(foot_x, foot_y + i, current_cube)
                else:
                    current_cube = Cube(
                        style=Style(bg_color=bg_color[1], border=side_style[3], al=alignment[3], font=font_style[2]),
                        value='N/A')
                    self.core.write_cube_to_book(foot_x, foot_y + i, current_cube)
        self._merge_body_cell(foot_y + 3, foot_x, foot_y, foot_x,
                              border_pattern[side_style[3]], fill_pattern[bg_color[4]],
                              font_pattern[font_style[2]], alignment_pattern[alignment[1]])

        # render global
        for i in range(self.header['scale'][1]):
            if i < 3:
                current_cube = Cube(
                    style=Style(bg_color=bg_color[4], border=side_style[1], al=alignment[1], font=font_style[2]),
                    value='Target From Global')
                self.core.write_cube_to_book(foot_x + 1, foot_y + i, current_cube)
            elif i < 18:
                current_cube = Cube(
                    style=Style(bg_color=bg_color[4], border=side_style[1], al=alignment[3], font=font_style[1]),
                    value='N/A')
                self.core.write_cube_to_book(foot_x + 1, foot_y + i, current_cube)
            else:
                current_cube = Cube(
                    style=Style(bg_color=bg_color[3], border=side_style[1], al=alignment[3], font=font_style[2]),
                    value=None)
                self.core.write_cube_to_book(foot_x + 1, foot_y + i, current_cube)
        self._merge_body_cell(foot_y + 3, foot_x + 1, foot_y, foot_x + 1,
                              border_pattern[side_style[1]], fill_pattern[bg_color[4]],
                              font_pattern[font_style[2]], alignment_pattern[alignment[1]])

    def render(self):
        self._render_title()
        self._render_header()
        self._render_body()
        self._adjust_column_width(16)
        self._render_footer()
        self._cal_end_point()
        self._add_table_border()


if __name__ == '__main__':
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    c = RenderCore(ws)
    table = MainTable(c, 'I am Title!', common_header_sgm_1[1], page2, 2, 2)
    table.render()
    table.core.ws.sheet_view.showGridLines = False
    wb.save('SGM.xlsx')
